import json
import os
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from nio_md_prep.agglomeration_audit import (
    _case_size_hint,
    audit_agglomerations,
    rescue_agglomerations,
    write_rescue_pool,
)


def _campaign(root: Path, name: str, progress: dict | None) -> Path:
    campaign = root / name
    campaign.mkdir()
    (campaign / "agglomeration_manifest.json").write_text(
        json.dumps({"status": "XTB_REQUIRED", "xtb": {"enabled": True}, "replicas": []}),
        encoding="utf-8",
    )
    (campaign / "xtb_cases.tsv").write_text("xtb/case\t0\t0\n", encoding="utf-8")
    if progress is not None:
        (campaign / "audit_xtb_runs.py").write_text(
            "#!/usr/bin/env python3\n"
            "import json\n"
            f"data = {progress!r}\n"
            "open('xtb_progress.json', 'w').write(json.dumps(data))\n",
            encoding="utf-8",
        )
    return campaign


def test_audit_agglomerations_consolidates_campaigns(tmp_path, capsys):
    _campaign(
        tmp_path,
        "Me4PACz_agglo",
        {
            "total": 2,
            "completed": 2,
            "pending": 0,
            "counts_by_status": {"COMPLETE": 2},
            "cases": [
                {"array_index": 0, "case": "xtb/case0", "status": "COMPLETE", "charge": "0", "uhf": "0", "detail": "done"},
                {"array_index": 1, "case": "xtb/case1", "status": "COMPLETE", "charge": "0", "uhf": "0", "detail": "done"},
            ],
        },
    )
    _campaign(
        tmp_path,
        "mixed_agglo",
        {
            "total": 3,
            "completed": 1,
            "pending": 2,
            "counts_by_status": {"COMPLETE": 1, "PENDING": 2},
            "cases": [
                {"array_index": 0, "case": "xtb/case0", "status": "COMPLETE", "charge": "0", "uhf": "0", "detail": "done"},
                {"array_index": 1, "case": "xtb/case1", "status": "PENDING", "charge": "0", "uhf": "0", "detail": "not started"},
                {"array_index": 2, "case": "xtb/case2", "status": "PENDING", "charge": "0", "uhf": "0", "detail": "not started"},
            ],
        },
    )
    report = audit_agglomerations(tmp_path)
    assert report["campaigns"] == 2
    assert report["cases"] == 5
    assert report["completed"] == 3
    assert report["pending"] == 2
    assert report["campaign_counts_by_status"] == {"COMPLETE": 1, "IN_PROGRESS": 1}
    assert (tmp_path / "agglomeration_xtb_audit.csv").is_file()
    assert (tmp_path / "agglomeration_xtb_audit.json").is_file()
    assert (tmp_path / "agglomeration_xtb_audit_cases.csv").is_file()
    assert (tmp_path / "agglomeration_xtb_audit_incomplete.csv").is_file()
    workbook_path = tmp_path / "agglomeration_xtb_audit.xlsx"
    assert workbook_path.is_file()
    assert load_workbook(workbook_path, read_only=True).sheetnames == [
        "Overview",
        "Campaigns",
        "Incomplete Cases",
        "Attention",
        "Pending",
        "All Cases",
        "Case Integrity",
        "Artifacts",
        "Logs",
        "Protocols",
        "Resources",
        "Slurm Jobs",
        "Status Counts",
        "Campaign Issues",
        "Scan Warnings",
        "Audit Metadata",
        "Status Definitions",
    ]
    assert len(report["incomplete_cases"]) == 2
    assert "Overall xTB progress: 3/5 safely complete" in capsys.readouterr().out


def test_case_diagnostics_identify_exact_checkpoint_and_log(tmp_path):
    campaign = _campaign(
        tmp_path,
        "diagnostic_agglo",
        {
            "total": 1,
            "completed": 0,
            "pending": 1,
            "counts_by_status": {"STEERING_COMPLETE": 1},
            "cases": [
                {
                    "array_index": 4,
                    "case": "xtb/n08/r000_s00_1p000",
                    "status": "STEERING_COMPLETE",
                    "charge": "0",
                    "uhf": "0",
                    "detail": "unbiased MD and final optimization remain",
                }
            ],
        },
    )
    work = campaign / "xtb/n08/r000_s00_1p000"
    work.mkdir(parents=True)
    (work / "xtbopt_initial.xyz").write_text("initial\n", encoding="utf-8")
    (work / "xtbsteered_last.xyz").write_text("steered\n", encoding="utf-8")
    (work / "xtb_unbiased_md.err").write_text(
        "DUE TO TIME LIMIT\n", encoding="utf-8"
    )
    (campaign / "xtb-pool.12345.out").write_text(
        ">>> Starting xTB case index 4 at Tue Aug 26 12:00:00 UTC 2026\n",
        encoding="utf-8",
    )
    (campaign / "xtb-pool.12345.err").write_text(
        "<<< Finished xTB case index 4 with ERROR 1 at Tue Aug 26 12:10:00 UTC 2026\n",
        encoding="utf-8",
    )
    report = audit_agglomerations(tmp_path, include_slurm=False)
    case = report["incomplete_cases"][0]
    assert case["campaign"] == "diagnostic_agglo"
    assert case["array_index"] == 4
    assert case["case"] == "xtb/n08/r000_s00_1p000"
    assert case["initial_opt_complete"] is True
    assert case["steered_md_complete"] is True
    assert case["unbiased_md_complete"] is False
    assert case["likely_issue"] == "WALLTIME_EXCEEDED"
    assert case["latest_log"] == "xtb_unbiased_md.err"
    assert case["latest_job_id"] == "12345"
    assert case["worker_state"] == "FAILED"
    assert case["worker_exit_code"] == "1"


def test_audit_agglomerations_flags_missing_auditor(tmp_path):
    _campaign(tmp_path, "legacy_agglo", None)
    report = audit_agglomerations(tmp_path)
    row = report["campaign_results"][0]
    assert row["campaign_status"] == "ERROR"
    assert "regenerate" in row["detail"]


def test_audit_agglomerations_flags_agglo_folder_without_manifest(tmp_path):
    (tmp_path / "broken_agglo").mkdir()
    report = audit_agglomerations(tmp_path)
    row = report["campaign_results"][0]
    assert row["campaign_status"] == "ERROR"
    assert row["detail"] == "manifest is missing"


def test_audit_agglomerations_detects_self_inconsistent_progress(tmp_path):
    _campaign(
        tmp_path,
        "corrupt_agglo",
        {
            "total": 99,
            "completed": 99,
            "pending": 0,
            "counts_by_status": {"COMPLETE": 99},
            "cases": [
                {
                    "array_index": 0,
                    "case": "xtb/case0",
                    "status": "PENDING",
                    "charge": "0",
                    "uhf": "0",
                    "detail": "not started",
                }
            ],
        },
    )
    row = audit_agglomerations(tmp_path)["campaign_results"][0]
    assert row["campaign_status"] == "ATTENTION"
    assert row["data_consistent"] is False
    assert "inconsistent audit output" in row["detail"]


def test_audit_agglomerations_strict_exit_code_via_cli(tmp_path, monkeypatch, capsys):
    from nio_md_prep import cli

    _campaign(tmp_path, "legacy_agglo", None)
    monkeypatch.chdir(tmp_path)
    assert cli.main(["audit-agglomerations", "--strict", "--no-slurm"]) == 1
    capsys.readouterr()


def _stuck_campaign(root: Path, name: str, cases: list[str]) -> Path:
    """A campaign whose every case is stuck mid-protocol (no error), like a
    carousel wall-clock timeout would leave behind."""
    campaign = _campaign(
        root,
        name,
        {
            "total": len(cases),
            "completed": 0,
            "pending": len(cases),
            "counts_by_status": {"STEERING_COMPLETE": len(cases)},
            "cases": [
                {
                    "array_index": index,
                    "case": case,
                    "status": "STEERING_COMPLETE",
                    "charge": "0",
                    "uhf": "0",
                    "detail": "unbiased MD and final optimization remain",
                }
                for index, case in enumerate(cases)
            ],
        },
    )
    (campaign / "run_xtb_array.sbatch").write_text("#!/bin/bash\necho fake\n", encoding="utf-8")
    return campaign


def test_case_size_hint_orders_largest_aggregate_first():
    assert _case_size_hint("xtb/n08/r000_s00_1p000") == 8
    assert _case_size_hint("xtb/n02/r000_s00_1p000") == 2
    assert _case_size_hint("xtb/mixed-2-2/r000_s00_1p000") == 4
    assert _case_size_hint("xtb/unrecognized-name/r000") == 0


def test_rescue_agglomerations_orders_largest_aggregate_first_across_campaigns(tmp_path, capsys):
    _stuck_campaign(
        tmp_path, "DCZ4P_agglo",
        ["xtb/n06/r000_s00_1p000", "xtb/n06/r001_s00_1p000", "xtb/n08/r000_s00_1p000"],
    )
    _stuck_campaign(tmp_path, "MeO-4PADBC_agglo", ["xtb/n08/r000_s00_1p000"])
    capsys.readouterr()

    result = rescue_agglomerations(tmp_path, dry_run=True)
    capsys.readouterr()

    # n08 (size 8) cases must be queued ahead of every n06 (size 6) case,
    # across both campaigns -- this is the fix for the largest aggregates
    # starting last behind smaller, already-finished cases in a shared
    # single-campaign carousel.
    ordered_sizes = [_case_size_hint(str(c["case"])) for c in result["cases"]]
    assert ordered_sizes == sorted(ordered_sizes, reverse=True)
    assert ordered_sizes[0] == 8 and ordered_sizes[1] == 8

    # 4 cases fit one node of 8 workers x 8 cores with nothing wasted.
    assert result["workers_per_node"] == 8
    assert result["nodes_needed"] == 1
    assert result["row_chunks"] == ["0-3"]

    cases_tsv = Path(result["cases_tsv"])
    rows = cases_tsv.read_text(encoding="utf-8").splitlines()
    assert len(rows) == 4
    first_campaign_root, first_index, first_case, first_label = rows[0].split("\t")
    assert first_case == "xtb/n08/r000_s00_1p000"
    assert Path(first_campaign_root).is_dir()

    pool_script = Path(result["pool_script"])
    subprocess.run(["bash", "-n", str(pool_script)], check=True)
    script_text = pool_script.read_text(encoding="utf-8")
    assert "#SBATCH --ntasks=8" in script_text
    assert "#SBATCH --cpus-per-task=8" in script_text
    assert "#SBATCH --exclusive" in script_text
    assert "#SBATCH -t 72:00:00" in script_text


def test_write_rescue_pool_rejects_node_cpus_not_divisible_by_cpus(tmp_path):
    try:
        write_rescue_pool(
            [{"campaign_path": str(tmp_path), "array_index": 0, "case": "xtb/n02/r0", "campaign": "x"}],
            tmp_path / "out",
            cpus=7,
            node_cpus=64,
        )
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "multiple" in str(exc)


def test_rescue_agglomerations_submits_one_job_per_node_chunk(tmp_path, monkeypatch):
    _stuck_campaign(
        tmp_path, "DCZ4P_agglo",
        ["xtb/n06/r000_s00_1p000", "xtb/n06/r001_s00_1p000", "xtb/n08/r000_s00_1p000"],
    )

    fake_bin = tmp_path / "fake-bin"
    fake_bin.mkdir()
    fake_sbatch = fake_bin / "sbatch"
    fake_sbatch.write_text(
        "#!/bin/bash\nprintf '%s\\n' \"$*\" >> \"$SBATCH_LOG\"\nprintf 'Submitted batch job 42\\n'\n",
        encoding="utf-8",
    )
    fake_sbatch.chmod(0o755)
    sbatch_log = tmp_path / "sbatch.log"
    monkeypatch.setenv("PATH", f"{fake_bin}{os.pathsep}{os.environ['PATH']}")
    monkeypatch.setenv("SBATCH_LOG", str(sbatch_log))

    # cpus=16 -> 4 workers/node, so the 3 rescued cases still fit in one
    # node/chunk (0-2); this exercises the actual submission path (sbatch
    # invoked from the rescue output directory, one call per node chunk).
    result = rescue_agglomerations(
        tmp_path,
        cpus=16,
        assume_yes=True,
        output_dir=tmp_path / "rescue_xtb",
    )
    assert result["workers_per_node"] == 4
    assert result["row_chunks"] == ["0-2"]
    assert len(result["submitted_jobs"]) == 1
    logged = sbatch_log.read_text(encoding="utf-8").strip()
    assert "--export=ALL,RESCUE_ROWS=0-2" in logged
    assert "rescue_xtb_pool.sbatch" in logged
