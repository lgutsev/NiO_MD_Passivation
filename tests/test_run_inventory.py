from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "inventory_runs", ROOT / "scripts/inventory_runs.py"
)
assert SPEC and SPEC.loader
inventory = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = inventory
SPEC.loader.exec_module(inventory)


def write(path: Path, text: str = "x\n") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def test_inventory_distinguishes_complete_recoverable_and_missing(tmp_path):
    repo = tmp_path
    (repo / ".git").mkdir()
    prepared = repo / "prepared"

    complete = prepared / "me-4pacz-then-meo-2pacz"
    write(complete / "topology_output.lmp")
    write(complete / "deposition.in")
    write(complete / "deposited.data")
    write(complete / "hold-300K.in", "run 1000\n")
    write(complete / "held-300K.data")

    recoverable = prepared / "me-4pacz-then-dcz-4p"
    write(recoverable / "topology_output.lmp")
    write(recoverable / "deposition.in")
    write(recoverable / "deposited.data")
    write(recoverable / "hold-300K.in", "run 1000\n")
    write(
        recoverable / "log.hold-300K.123.lammps",
        "Loop time of 10 on 64 procs for 1000 steps with 100 atoms\n",
    )
    write(recoverable / "hold-300K.restart.1")

    runs = inventory.discover_runs([prepared])
    records = [
        record
        for directory, root in runs
        for record in inventory.inventory_run(directory, root, repo)
    ]
    lookup = {(row.system, row.stage): row.status for row in records}

    assert lookup[("me-4pacz-then-meo-2pacz", "Hold 300 K")] == "COMPLETE"
    assert (
        lookup[("me-4pacz-then-meo-2pacz", "LEGO continuation")]
        == "NOT_APPLICABLE"
    )
    assert (
        lookup[("me-4pacz-then-dcz-4p", "Hold 300 K")]
        == "FINALIZE_NEEDED"
    )
    assert (
        lookup[("me-4pacz-then-dcz-4p", "DCZ control nonsticky 300K")]
        == "BLOCKED"
    )


def test_inventory_workbook_contains_auditable_sheets(tmp_path):
    repo = tmp_path
    (repo / ".git").mkdir()
    prepared = repo / "prepared-lego"
    run = prepared / "me-4pacz-then-dcz-4p-lego-seeded"
    write(run / "topology_output.lmp")
    write(run / "deposition.in")
    write(run / "deposited.data")
    write(run / "continue-deposition.in")
    write(run / "deposition-continuation.complete")
    write(run / "hold-300K.in")

    records = inventory.inventory_run(run, prepared, repo)
    output = repo / "inventory.xlsx"
    inventory.build_workbook(
        repo, [prepared], records, output, include_slurm=False
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Dashboard",
        "Run Status",
        "Stage Detail",
        "Action Queue",
        "Analysis Reports",
        "Current Slurm Jobs",
        "README",
    ]
    assert workbook["Dashboard"]["B7"].value == 1
    assert workbook["Run Status"]["A2"].value == "prepared-lego"
    assert workbook["Stage Detail"].max_row > 10
    assert workbook["Action Queue"].max_row > 1
