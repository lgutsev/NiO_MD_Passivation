#!/usr/bin/env python3
"""Inventory every NiO prepared tree and write an audit-friendly XLSX report."""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - exercised on QBD only
    raise SystemExit(
        "openpyxl is required. Activate .venv and run "
        "python -m pip install -e '.[analysis]'."
    ) from exc


STATUS_COLORS = {
    "COMPLETE": "C6EFCE",
    "READY": "FFF2CC",
    "FINALIZE_NEEDED": "FCE4D6",
    "PARTIAL": "F4B183",
    "INPUT_MISSING": "F8CBAD",
    "BLOCKED": "E7E6E6",
    "ZERO_LENGTH": "FFC7CE",
    "NOT_APPLICABLE": "EDEDED",
    "NEEDS_ATTENTION": "FFC7CE",
}
STATUS_PRIORITY = {
    "ZERO_LENGTH": "Critical",
    "FINALIZE_NEEDED": "High",
    "PARTIAL": "High",
    "INPUT_MISSING": "Medium",
    "READY": "Medium",
    "BLOCKED": "Low",
}


@dataclass(frozen=True)
class StageSpec:
    name: str
    source: str | None
    input_file: str | None
    output: str
    trajectory: str | None = None
    restart_glob: str | None = None
    log_glob: str | None = None
    optional: bool = False
    analysis: bool = False


@dataclass
class StageRecord:
    prepared_root: str
    relative_run: str
    system: str
    family: str
    stage: str
    status: str
    source: str
    input_file: str
    output: str
    output_size_mb: float | None
    output_modified: datetime | None
    evidence: str
    recommended_action: str


STAGES = [
    StageSpec("Build", None, None, "topology_output.lmp"),
    StageSpec(
        "Deposition",
        "topology_output.lmp",
        "deposition.in",
        "deposited.data",
        "deposition.lammpstrj",
        "deposition.restart.*",
        "log.deposition.*.lammps",
    ),
    StageSpec(
        "LEGO continuation",
        "deposited.data",
        "continue-deposition.in",
        "deposition-continuation.complete",
        "deposition-continuation.lammpstrj",
        "deposition-continuation.restart.*",
        "log.deposition-continuation.*.lammps",
        optional=True,
    ),
    StageSpec(
        "Hold 300 K",
        "deposited.data",
        "hold-300K.in",
        "held-300K.data",
        "hold-300K.lammpstrj",
        "hold-300K.restart.*",
        "log.hold-300K.*.lammps",
    ),
    StageSpec(
        "Hold 400 K",
        "deposited.data",
        "hold-400K.in",
        "held-400K.data",
        "hold-400K.lammpstrj",
        "hold-400K.restart.*",
        "log.hold-400K.*.lammps",
    ),
    StageSpec(
        "Relaxed 300 K",
        "held-300K.data",
        "decompress-300K.in",
        "relaxed-300K.data",
        "relax-300K.lammpstrj",
        "decompress-300K.restart.*",
        "log.decompress-300K.*.lammps",
    ),
    StageSpec(
        "Relaxed 400 K",
        "held-400K.data",
        "decompress-400K.in",
        "relaxed-400K.data",
        "relax-400K.lammpstrj",
        "decompress-400K.restart.*",
        "log.decompress-400K.*.lammps",
    ),
]

CONTROL_STAGES = [
    StageSpec(
        f"DCZ control {mode} {temperature}",
        f"held-{temperature}.data",
        f"decompress-control-{mode}-{temperature}.in",
        f"relaxed-control-{mode}-{temperature}.data",
        f"relax-control-{mode}-{temperature}.lammpstrj",
        f"relax-control-{mode}-{temperature}.restart.*",
        f"log.decompress-control-{mode}-{temperature}.*.lammps",
        optional=True,
    )
    for temperature in ("300K", "400K")
    for mode in ("nonsticky", "nowall")
]

ANALYSIS_STAGES = [
    StageSpec(
        f"Coverage {label}",
        trajectory,
        None,
        f"coverage-analysis-{label}/coverage_summary.json",
        optional=True,
        analysis=True,
    )
    for label, trajectory in (
        ("hold-300K", "hold-300K.lammpstrj"),
        ("hold-400K", "hold-400K.lammpstrj"),
        ("relax-300K", "relax-300K.lammpstrj"),
        ("relax-400K", "relax-400K.lammpstrj"),
    )
] + [
    StageSpec(
        f"Interface {label}",
        trajectory,
        None,
        f"interface-analysis-{label}/interface_summary.json",
        optional=True,
        analysis=True,
    )
    for label, trajectory in (
        ("deposition", "deposition.lammpstrj"),
        ("hold-300K", "hold-300K.lammpstrj"),
        ("hold-400K", "hold-400K.lammpstrj"),
        ("relax-300K", "relax-300K.lammpstrj"),
        ("relax-400K", "relax-400K.lammpstrj"),
    )
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument(
        "--prepared-root",
        action="append",
        type=Path,
        default=[],
        help="Prepared tree to scan; repeat as needed. Defaults to repo/prepared*.",
    )
    parser.add_argument(
        "--output", type=Path, default=Path("run_inventory.xlsx")
    )
    parser.add_argument("--no-slurm", action="store_true")
    return parser.parse_args()


def nonempty(path: Path) -> bool:
    return path.is_file() and path.stat().st_size > 0


def newest(paths: Iterable[Path]) -> Path | None:
    candidates = [path for path in paths if path.is_file()]
    return max(candidates, key=lambda path: path.stat().st_mtime, default=None)


def configured_final_run(input_path: Path | None) -> str | None:
    if input_path is None or not nonempty(input_path):
        return None
    expected = None
    for line in input_path.read_text(
        encoding="utf-8", errors="replace"
    ).splitlines():
        fields = line.split()
        if fields and fields[0] == "run" and len(fields) >= 2:
            token = fields[1]
            if token.isdigit():
                expected = token
    return expected


def completed_loop(log_path: Path | None, input_path: Path | None) -> bool:
    expected = configured_final_run(input_path)
    if log_path is None or expected is None:
        return False
    text = log_path.read_text(encoding="utf-8", errors="replace")
    return re.search(
        rf"Loop time.*for\s+{re.escape(expected)}\s+steps", text
    ) is not None


def family_for(directory: Path) -> str:
    name = directory.name
    if name.endswith("-lego2-seeded"):
        return "LEGO2"
    if name.endswith("-lego-seeded"):
        return "LEGO1"
    if "replicates" in directory.parts or re.search(r"seed-\d+", str(directory)):
        return "Replicate"
    if "-then-" in name:
        return "Sequential"
    if "-cosam" in name:
        return "CoSAM"
    return "Primary"


def discover_runs(prepared_roots: list[Path]) -> list[tuple[Path, Path]]:
    found: dict[Path, Path] = {}
    for prepared_root in prepared_roots:
        if not prepared_root.is_dir():
            continue
        for marker in ("topology_output.lmp", "assembly_manifest.json", "deposition.in"):
            for path in prepared_root.rglob(marker):
                directory = path.parent.resolve()
                found.setdefault(directory, prepared_root.resolve())
    return sorted(found.items(), key=lambda item: str(item[0]))


def evidence_for(
    directory: Path, spec: StageSpec
) -> tuple[str, list[Path], Path | None]:
    pieces: list[str] = []
    partial_files: list[Path] = []
    if spec.trajectory:
        trajectory = directory / spec.trajectory
        if trajectory.exists():
            partial_files.append(trajectory)
            pieces.append(
                f"trajectory={trajectory.name} "
                f"({trajectory.stat().st_size / 2**20:.1f} MB)"
            )
    if spec.restart_glob:
        restarts = list(directory.glob(spec.restart_glob))
        if restarts:
            partial_files.extend(restarts)
            pieces.append(f"restarts={len(restarts)}")
    latest_log = newest(directory.glob(spec.log_glob)) if spec.log_glob else None
    if latest_log:
        partial_files.append(latest_log)
        pieces.append(f"latest_log={latest_log.name}")
    return "; ".join(pieces), partial_files, latest_log


def action_for(spec: StageSpec, status: str) -> str:
    if status in {"COMPLETE", "NOT_APPLICABLE"}:
        return ""
    if status == "ZERO_LENGTH":
        return f"Archive the zero-length {spec.output}, then rerun or recover."
    if status == "FINALIZE_NEEDED":
        return "Recover held data from the final alternating restart."
    if status == "PARTIAL":
        return "Inspect the latest log, archive stale outputs, and relaunch."
    if status == "INPUT_MISSING":
        return "Regenerate stage inputs without rebuilding topology."
    if status == "READY":
        return "Submit this stage."
    return f"Complete prerequisite: {spec.source}." if spec.source else "Build the system."


def stage_record(
    directory: Path,
    prepared_root: Path,
    repo: Path,
    spec: StageSpec,
    expected: bool,
) -> StageRecord:
    output = directory / spec.output
    source = directory / spec.source if spec.source else None
    input_path = directory / spec.input_file if spec.input_file else None
    evidence, partial_files, latest_log = evidence_for(directory, spec)

    if not expected:
        status = "NOT_APPLICABLE"
    elif output.exists() and output.stat().st_size == 0:
        status = "ZERO_LENGTH"
    elif nonempty(output):
        status = "COMPLETE"
    elif (
        spec.name.startswith("Hold ")
        and completed_loop(latest_log, input_path)
        and any("restart" in path.name for path in partial_files)
    ):
        status = "FINALIZE_NEEDED"
    elif partial_files:
        status = "PARTIAL"
    elif source is not None and not nonempty(source):
        status = "BLOCKED"
    elif input_path is not None and not nonempty(input_path):
        status = "INPUT_MISSING"
    else:
        status = "READY"

    stat = output.stat() if output.exists() else None
    return StageRecord(
        prepared_root=str(prepared_root.relative_to(repo)),
        relative_run=str(directory.relative_to(repo)),
        system=directory.name,
        family=family_for(directory),
        stage=spec.name,
        status=status,
        source=spec.source or "",
        input_file=spec.input_file or "",
        output=spec.output,
        output_size_mb=round(stat.st_size / 2**20, 3) if stat else None,
        output_modified=(
            datetime.fromtimestamp(stat.st_mtime).astimezone().replace(tzinfo=None)
            if stat
            else None
        ),
        evidence=evidence,
        recommended_action=action_for(spec, status),
    )


def inventory_run(
    directory: Path, prepared_root: Path, repo: Path
) -> list[StageRecord]:
    records = [
        stage_record(
            directory,
            prepared_root,
            repo,
            spec,
            not spec.optional
            or family_for(directory) in {"LEGO1", "LEGO2"}
            or (
                spec.input_file is not None
                and (directory / spec.input_file).exists()
            )
            or (directory / spec.output).exists(),
        )
        for spec in STAGES
    ]
    is_dcz_control_target = (
        directory.parent.resolve() == prepared_root.resolve()
        and (
            (
                prepared_root.name
                in {"prepared", "prepared-rerun", "prepared-rerun2"}
                and directory.name == "me-4pacz-then-dcz-4p"
            )
            or (
                prepared_root.name == "prepared-lego"
                and directory.name == "me-4pacz-then-dcz-4p-lego-seeded"
            )
        )
    )
    records.extend(
        stage_record(
            directory,
            prepared_root,
            repo,
            spec,
            is_dcz_control_target
            or (directory / str(spec.input_file)).exists()
            or (directory / spec.output).exists(),
        )
        for spec in CONTROL_STAGES
    )
    records.extend(
        stage_record(
            directory,
            prepared_root,
            repo,
            spec,
            (directory / str(spec.source)).exists()
            or (directory / spec.output).exists(),
        )
        for spec in ANALYSIS_STAGES
    )
    return records


def run_rows(records: list[StageRecord]) -> list[list[object]]:
    grouped: dict[str, list[StageRecord]] = {}
    for record in records:
        grouped.setdefault(record.relative_run, []).append(record)
    rows = []
    standard_order = [spec.name for spec in STAGES]
    for path, group in sorted(grouped.items()):
        relevant = [row for row in group if row.status != "NOT_APPLICABLE"]
        attention = [
            row
            for row in relevant
            if row.status
            in {"ZERO_LENGTH", "FINALIZE_NEEDED", "PARTIAL", "INPUT_MISSING"}
        ]
        ready = [row for row in relevant if row.status == "READY"]
        blocked = [row for row in relevant if row.status == "BLOCKED"]
        overall = (
            "NEEDS_ATTENTION"
            if attention
            else "READY"
            if ready
            else "BLOCKED"
            if blocked
            else "COMPLETE"
        )
        latest = next(
            (
                name
                for name in reversed(standard_order)
                if any(row.stage == name and row.status == "COMPLETE" for row in group)
            ),
            "None",
        )
        first_action = next(
            (
                row.recommended_action
                for row in relevant
                if row.recommended_action
            ),
            "",
        )
        exemplar = group[0]
        rows.append(
            [
                exemplar.prepared_root,
                path,
                exemplar.system,
                exemplar.family,
                overall,
                latest,
                len(attention),
                len(ready),
                len(blocked),
                sum(row.status == "COMPLETE" for row in relevant),
                len(relevant),
                first_action,
            ]
        )
    return rows


def slurm_rows() -> list[list[str]]:
    try:
        result = subprocess.run(
            [
                "squeue",
                "-h",
                "-u",
                subprocess.check_output(["id", "-un"], text=True).strip(),
                "-o",
                "%i|%j|%T|%M|%l|%D|%C|%R",
            ],
            check=False,
            capture_output=True,
            text=True,
        )
    except (FileNotFoundError, subprocess.SubprocessError):
        return []
    if result.returncode:
        return []
    return [line.split("|", 7) for line in result.stdout.splitlines() if line.strip()]


def report_rows(prepared_roots: list[Path], repo: Path) -> list[list[object]]:
    rows = []
    for root in prepared_roots:
        for name in (
            "coverage_summary.xlsx",
            "interface_structure_summary.xlsx",
            "publication_summary.xlsx",
        ):
            path = root / name
            stat = path.stat() if path.exists() else None
            detail_count = (
                sum(1 for _ in root.rglob("coverage_summary.json"))
                if name == "coverage_summary.xlsx"
                else sum(1 for _ in root.rglob("interface_summary.json"))
                if name == "interface_structure_summary.xlsx"
                else None
            )
            rows.append(
                [
                    str(root.relative_to(repo)),
                    name,
                    "COMPLETE" if stat and stat.st_size else "MISSING",
                    round(stat.st_size / 2**20, 3) if stat else None,
                    (
                        datetime.fromtimestamp(stat.st_mtime)
                        .astimezone()
                        .replace(tzinfo=None)
                        if stat
                        else None
                    ),
                    detail_count,
                ]
            )
    return rows


def add_table_sheet(
    workbook: Workbook,
    name: str,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
):
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if rows:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    return sheet


def style_sheet(
    sheet, widths: dict[str, float], status_column: str | None = None
) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    sheet.row_dimensions[1].height = 30
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
    if status_column and sheet.max_row >= 2:
        for status, color in STATUS_COLORS.items():
            sheet.conditional_formatting.add(
                f"{status_column}2:{status_column}{sheet.max_row}",
                FormulaRule(
                    formula=[f'${status_column}2="{status}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )


def build_workbook(
    repo: Path,
    prepared_roots: list[Path],
    records: list[StageRecord],
    output: Path,
    include_slurm: bool,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    runs = run_rows(records)
    stages = [
        [
            row.prepared_root,
            row.relative_run,
            row.system,
            row.family,
            row.stage,
            row.status,
            row.source,
            row.input_file,
            row.output,
            row.output_size_mb,
            row.output_modified,
            row.evidence,
            row.recommended_action,
        ]
        for row in records
    ]
    actions = [
        [
            STATUS_PRIORITY.get(row.status, "Low"),
            row.status,
            row.prepared_root,
            row.relative_run,
            row.stage,
            row.recommended_action,
            row.evidence,
        ]
        for row in records
        if row.status in STATUS_PRIORITY
    ]
    priority_rank = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    actions.sort(
        key=lambda row: (
            priority_rank[str(row[0])],
            str(row[2]),
            str(row[3]),
            str(row[4]),
        )
    )
    slurm = slurm_rows() if include_slurm else []
    reports = report_rows(prepared_roots, repo)
    generated = datetime.now().astimezone().replace(tzinfo=None)

    dashboard = workbook.create_sheet("Dashboard")
    dashboard.sheet_view.showGridLines = False
    dashboard["A1"] = "NiO MD Campaign Inventory"
    dashboard["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    dashboard["A1"].fill = PatternFill("solid", fgColor="17365D")
    dashboard.merge_cells("A1:H2")
    dashboard["A1"].alignment = Alignment(vertical="center")
    dashboard["A4"], dashboard["B4"] = "Generated", generated
    dashboard["B4"].number_format = "yyyy-mm-dd hh:mm"
    dashboard["D4"], dashboard["E4"] = "Repository", str(repo)
    for cell, value in {
        "A6": "KPI",
        "B6": "Value",
        "D6": "Run state",
        "E6": "Count",
    }.items():
        dashboard[cell] = value
        dashboard[cell].fill = PatternFill("solid", fgColor="5B9BD5")
        dashboard[cell].font = Font(color="FFFFFF", bold=True)
    kpis = [
        ("Prepared roots", len(prepared_roots)),
        ("Simulation runs", len(runs)),
        ("Stage records", len(stages)),
        ("Action items", len(actions)),
        ("Current Slurm jobs", len(slurm)),
        ("Analysis workbooks present", sum(row[2] == "COMPLETE" for row in reports)),
    ]
    for index, row in enumerate(kpis, 7):
        dashboard.cell(index, 1, row[0])
        dashboard.cell(index, 2, row[1])
    for index, state in enumerate(
        ("COMPLETE", "READY", "BLOCKED", "NEEDS_ATTENTION"), 7
    ):
        dashboard.cell(index, 4, state)
        dashboard.cell(index, 5, sum(row[4] == state for row in runs))
    dashboard.freeze_panes = "A4"
    for column, width in {"A": 28, "B": 14, "C": 4, "D": 22, "E": 48}.items():
        dashboard.column_dimensions[column].width = width

    run_sheet = add_table_sheet(
        workbook,
        "Run Status",
        [
            "Prepared Root",
            "Relative Run",
            "System",
            "Family",
            "Overall Status",
            "Latest Complete Stage",
            "Attention Items",
            "Ready Stages",
            "Blocked Stages",
            "Completed Stages",
            "Expected Stages",
            "Next Action",
        ],
        runs,
        "RunStatusTable",
    )
    style_sheet(
        run_sheet,
        {
            "A": 20,
            "B": 52,
            "C": 42,
            "D": 14,
            "E": 20,
            "F": 24,
            "G": 14,
            "H": 13,
            "I": 14,
            "J": 16,
            "K": 15,
            "L": 58,
        },
        "E",
    )

    stage_sheet = add_table_sheet(
        workbook,
        "Stage Detail",
        [
            "Prepared Root",
            "Relative Run",
            "System",
            "Family",
            "Stage",
            "Status",
            "Required Source",
            "Input File",
            "Final Output",
            "Output Size (MB)",
            "Modified",
            "Evidence",
            "Recommended Action",
        ],
        stages,
        "StageDetailTable",
    )
    style_sheet(
        stage_sheet,
        {
            "A": 18,
            "B": 48,
            "C": 40,
            "D": 13,
            "E": 28,
            "F": 20,
            "G": 28,
            "H": 34,
            "I": 42,
            "J": 17,
            "K": 19,
            "L": 58,
            "M": 62,
        },
        "F",
    )
    for cell in stage_sheet["J"][1:]:
        cell.number_format = "0.000"

    action_sheet = add_table_sheet(
        workbook,
        "Action Queue",
        [
            "Priority",
            "Status",
            "Prepared Root",
            "Relative Run",
            "Stage",
            "Recommended Action",
            "Evidence",
        ],
        actions,
        "ActionQueueTable",
    )
    style_sheet(
        action_sheet,
        {"A": 12, "B": 20, "C": 18, "D": 50, "E": 28, "F": 65, "G": 60},
        "B",
    )

    report_sheet = add_table_sheet(
        workbook,
        "Analysis Reports",
        [
            "Prepared Root",
            "Workbook",
            "Status",
            "Size (MB)",
            "Modified",
            "Detailed JSON Summaries",
        ],
        reports,
        "AnalysisReportsTable",
    )
    style_sheet(
        report_sheet,
        {"A": 20, "B": 40, "C": 16, "D": 14, "E": 19, "F": 24},
        "C",
    )

    slurm_sheet = add_table_sheet(
        workbook,
        "Current Slurm Jobs",
        [
            "Job ID",
            "Job Name",
            "State",
            "Elapsed",
            "Time Limit",
            "Nodes",
            "CPUs",
            "Reason/Node",
        ],
        slurm,
        "CurrentSlurmJobsTable",
    )
    style_sheet(
        slurm_sheet,
        {"A": 18, "B": 30, "C": 15, "D": 14, "E": 14, "F": 10, "G": 10, "H": 42},
    )

    methods = workbook.create_sheet("README")
    methods.sheet_view.showGridLines = False
    methods.append(["Field", "Definition"])
    notes = [
        ("Purpose", "File-based inventory of every simulation under prepared* trees."),
        (
            "Authority",
            "Non-zero scientific output files determine completion; Slurm exit state alone does not.",
        ),
        (
            "FINALIZE_NEEDED",
            "The latest hold log completed its configured run and a restart exists, but held-*.data is absent.",
        ),
        (
            "PARTIAL",
            "Trajectory/restart/log evidence exists without the required final output.",
        ),
        ("READY", "Prerequisite and stage input exist; the stage can be submitted."),
        ("INPUT_MISSING", "Prerequisite exists but the generated input is absent."),
        ("BLOCKED", "The required upstream scientific output is absent."),
        (
            "Scope",
            "A run directory contains topology_output.lmp, assembly_manifest.json, or deposition.in.",
        ),
        (
            "Generated",
            generated.isoformat(sep=" ", timespec="seconds"),
        ),
    ]
    for row in notes:
        methods.append(row)
    style_sheet(methods, {"A": 24, "B": 110})

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    load_workbook(output, read_only=True).close()


def main() -> int:
    args = parse_args()
    repo = args.repo.resolve()
    if not (repo / ".git").exists():
        raise SystemExit(f"Not a Git repository root: {repo}")
    prepared_roots = (
        [path.resolve() for path in args.prepared_root]
        if args.prepared_root
        else sorted(path.resolve() for path in repo.glob("prepared*") if path.is_dir())
    )
    runs = discover_runs(prepared_roots)
    records = [
        record
        for directory, prepared_root in runs
        for record in inventory_run(directory, prepared_root, repo)
    ]
    output = args.output if args.output.is_absolute() else repo / args.output
    build_workbook(repo, prepared_roots, records, output, not args.no_slurm)
    print(f"Wrote {output}")
    print(f"Prepared roots: {len(prepared_roots)}")
    print(f"Simulation runs: {len(runs)}")
    print(f"Stage records: {len(records)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
