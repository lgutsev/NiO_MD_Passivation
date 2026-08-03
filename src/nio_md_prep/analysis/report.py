"""Consolidated Excel reporting for projected coverage analyses."""
from __future__ import annotations

from pathlib import Path
from typing import Any
import json
import math
import re


RESULT_HEADERS = [
    "Run",
    "System",
    "Temperature (K)",
    "Stage",
    "Frames Analyzed",
    "Total Coverage (%)",
    "Block SEM (%)",
    "Frame SD (%)",
    "Uncovered (%)",
    "Component Overlap (%)",
    "Void Patch Count",
    "Largest Void Patch (% of cell)",
    "Mean Void Patch (% of cell)",
    "Roughness RMS (Å)",
    "Roughness Mean Absolute (Å)",
    "Roughness Peak-to-Valley (Å)",
    "Primary Component",
    "Primary Coverage (%)",
    "Secondary Component",
    "Secondary Coverage (%)",
    "Grid Spacing (Å)",
    "Radius Scale",
    "Hydrogen Included",
    "QC Balance (%)",
    "QC Union (%)",
    "Status",
    "Trajectory",
    "Summary JSON",
    "Run Directory",
    "Near-Surface Coverage (%)",
    "P-Near-Surface-Conditioned Coverage (%)",
    "Near-Surface Largest Void Patch (% of cell)",
    "P-Near-Surface Largest Void Patch (% of cell)",
    "Scientific Scope",
]

COMPONENT_HEADERS = [
    "Run",
    "System",
    "Temperature (K)",
    "Component",
    "Molecule ID Start",
    "Molecule ID End",
    "Coverage (%)",
    "Block SEM (%)",
    "Frame SD (%)",
]


def _openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference, Series, ScatterChart
        from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(
            "coverage workbook creation requires openpyxl; "
            "install with: python -m pip install -e '.[analysis]'"
        ) from exc
    return {
        "Workbook": Workbook,
        "BarChart": BarChart,
        "ScatterChart": ScatterChart,
        "Reference": Reference,
        "Series": Series,
        "ColorScaleRule": ColorScaleRule,
        "FormulaRule": FormulaRule,
        "Alignment": Alignment,
        "Font": Font,
        "PatternFill": PatternFill,
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
    }


def _temperature(stage: str, trajectory: str) -> int | None:
    match = re.search(r"(?:hold|relax)-(\d+)K", stage)
    if match is None:
        match = re.search(r"(?:hold|relax)-(\d+)K", Path(trajectory).name)
    return int(match.group(1)) if match else None


def _metric(summary: dict[str, Any], name: str, field: str) -> float:
    return float(summary["metrics"][name][field])


def collect_coverage_results(prepared_root: Path) -> tuple[list[dict], list[dict]]:
    """Collect one result row per completed hold and normalized component rows."""
    prepared_root = Path(prepared_root)
    paths = sorted(
        set(
            prepared_root.rglob(
                "coverage-analysis-hold-*/coverage_summary.json"
            )
        )
        | set(
            prepared_root.rglob(
                "coverage-analysis-relax-*/coverage_summary.json"
            )
        )
    )
    if not paths:
        raise FileNotFoundError(
            f"{prepared_root}: no compressed-hold or relaxed-hold coverage summaries "
            "files were found"
        )

    results: list[dict] = []
    component_rows: list[dict] = []
    for summary_path in paths:
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        run_directory_path = summary_path.parents[1]
        system = run_directory_path.name
        run_directory = str(run_directory_path.relative_to(prepared_root))
        stage = summary_path.parent.name.removeprefix("coverage-analysis-")
        trajectory = str(summary.get("trajectory", ""))
        temperature = _temperature(stage, trajectory)
        run = (
            f"{run_directory} | {temperature} K"
            if temperature is not None
            else f"{run_directory} | {stage}"
        )

        components = []
        for name, values in summary.get("components", {}).items():
            molecule_range = values.get("molecule_id_range", [None, None])
            lo = molecule_range[0] if len(molecule_range) > 0 else None
            hi = molecule_range[1] if len(molecule_range) > 1 else None
            components.append(
                {
                    "name": name,
                    "lo": lo,
                    "hi": hi,
                    "mean": float(values["mean_percent"]),
                    "sem": float(values["block_sem_percent"]),
                    "std": float(values["frame_std_percent"]),
                }
            )
        components.sort(
            key=lambda item: (
                math.inf if item["lo"] is None else int(item["lo"]),
                item["name"],
            )
        )

        total = _metric(summary, "total", "mean_percent")
        uncovered = _metric(summary, "uncovered", "mean_percent")
        overlap = _metric(summary, "overlap", "mean_percent")
        balance_qc = total + uncovered - 100.0
        union_qc = None
        if len(components) <= 2:
            union_qc = total - sum(item["mean"] for item in components) + overlap
        status = (
            "OK"
            if abs(balance_qc) <= 0.05
            and (union_qc is None or abs(union_qc) <= 0.05)
            else "CHECK"
        )

        primary = components[0] if components else None
        secondary = components[1] if len(components) > 1 else None
        try:
            relative_summary = str(summary_path.relative_to(prepared_root.parent))
        except ValueError:
            relative_summary = str(summary_path)

        results.append(
            {
                "run": run,
                "system": system,
                "temperature": temperature,
                "stage": stage,
                "frames": int(summary["frames_analyzed"]),
                "total": total,
                "sem": _metric(summary, "total", "block_sem_percent"),
                "std": _metric(summary, "total", "frame_std_percent"),
                "uncovered": uncovered,
                "overlap": overlap,
                "primary_name": primary["name"] if primary else None,
                "primary_mean": primary["mean"] if primary else None,
                "secondary_name": secondary["name"] if secondary else None,
                "secondary_mean": secondary["mean"] if secondary else None,
                # .get(...) rather than _metric(...): older coverage_summary.json
                # files predate these keys and must still load without error.
                "void_patch_count": (
                    summary.get("metrics", {}).get("void_patch_count", {}).get("mean")
                ),
                "void_largest_patch": (
                    summary.get("metrics", {})
                    .get("void_largest_patch_percent", {})
                    .get("mean_percent")
                ),
                "void_mean_patch": (
                    summary.get("metrics", {})
                    .get("void_mean_patch_percent", {})
                    .get("mean_percent")
                ),
                "roughness_rms": (
                    summary.get("metrics", {})
                    .get("roughness_rms", {})
                    .get("mean_angstrom")
                ),
                "roughness_mean_absolute": (
                    summary.get("metrics", {})
                    .get("roughness_mean_absolute", {})
                    .get("mean_angstrom")
                ),
                "roughness_peak_to_valley": (
                    summary.get("metrics", {})
                    .get("roughness_peak_to_valley", {})
                    .get("mean_angstrom")
                ),
                "near_surface": (
                    summary.get("metrics", {})
                    .get("near_surface", {})
                    .get("mean_percent")
                ),
                "anchor_conditioned": (
                    summary.get("metrics", {})
                    .get(
                        "p_near_surface_conditioned",
                        summary.get("metrics", {}).get("anchor_conditioned", {}),
                    )
                    .get("mean_percent")
                ),
                "near_surface_void_largest_patch": (
                    summary.get("metrics", {})
                    .get("near_surface_void_largest_patch_percent", {})
                    .get("mean_percent")
                ),
                "p_near_surface_void_largest_patch": (
                    summary.get("metrics", {})
                    .get("p_near_surface_void_largest_patch_percent", {})
                    .get("mean_percent")
                ),
                # Internal-only fields (not surfaced as workbook columns): used
                # by publication_report.py's comparability gate to detect a
                # coverage/interface pair analyzed over mismatched windows.
                "first_step": summary.get("provenance", {}).get("first_step"),
                "last_step": summary.get("provenance", {}).get("last_step"),
                "stride": summary.get("provenance", {}).get("stride"),
                "git_commit": summary.get("provenance", {}).get("git_commit"),
                "schema_version": summary.get("provenance", {}).get("schema_version"),
                "trajectory_sha256": summary.get("provenance", {}).get("trajectory_sha256"),
                "trajectory_size_bytes": summary.get("provenance", {}).get("trajectory_size_bytes"),
                "trajectory_mtime_utc": summary.get("provenance", {}).get("trajectory_mtime_utc"),
                "topology_sha256": summary.get("provenance", {}).get("topology_sha256"),
                "manifest_sha256": summary.get("provenance", {}).get("manifest_sha256"),
                "requested_blocks": summary.get("provenance", {}).get("requested_blocks"),
                "grid": float(summary["grid_target_spacing_angstrom"]),
                "radius_scale": float(summary["radius_scale"]),
                "hydrogen_included": not bool(summary["exclude_hydrogen"]),
                "near_surface_height": summary.get("near_surface_height_angstrom"),
                "surface_reference_depth": summary.get("surface_reference_depth_angstrom"),
                "balance_qc": balance_qc,
                "union_qc": union_qc,
                "status": status,
                "trajectory": trajectory,
                "summary_path": relative_summary,
                "method": str(summary["method"]),
                "run_directory": run_directory,
            }
        )
        for component in components:
            component_rows.append(
                {
                    "run": run,
                    "system": system,
                    "temperature": temperature,
                    "name": component["name"],
                    "lo": component["lo"],
                    "hi": component["hi"],
                    "mean": component["mean"],
                    "sem": component["sem"],
                    "std": component["std"],
                }
            )

    results.sort(
        key=lambda row: (
            row["system"],
            math.inf if row["temperature"] is None else row["temperature"],
            row["stage"],
        )
    )
    component_rows.sort(
        key=lambda row: (
            row["system"],
            math.inf if row["temperature"] is None else row["temperature"],
            math.inf if row["lo"] is None else row["lo"],
        )
    )
    return results, component_rows


def _style_table_sheet(sheet, headers: list[str], rows: list[list], table_name: str, api):
    Alignment = api["Alignment"]
    Font = api["Font"]
    PatternFill = api["PatternFill"]
    Table = api["Table"]
    TableStyleInfo = api["TableStyleInfo"]

    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "C2"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 32
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if rows:
        table = Table(displayName=table_name, ref=f"A1:{sheet.cell(1, len(headers)).column_letter}{len(rows)+1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def create_coverage_workbook(prepared_root: Path, output: Path) -> Path:
    """Create a formatted workbook summarizing all completed hold analyses."""
    api = _openpyxl()
    Workbook = api["Workbook"]
    BarChart = api["BarChart"]
    Reference = api["Reference"]
    ColorScaleRule = api["ColorScaleRule"]
    FormulaRule = api["FormulaRule"]
    PatternFill = api["PatternFill"]
    Font = api["Font"]

    results, components = collect_coverage_results(prepared_root)
    workbook = Workbook()
    workbook.properties.creator = "nio-md-prep"
    workbook.properties.title = "NiO passivant projected coverage summary"
    results_sheet = workbook.active
    results_sheet.title = "Results"
    component_sheet = workbook.create_sheet("Components")
    methods_sheet = workbook.create_sheet("Methods")

    result_rows = [
        [
            row["run"],
            row["system"],
            row["temperature"],
            row["stage"],
            row["frames"],
            row["total"],
            row["sem"],
            row["std"],
            row["uncovered"],
            row["overlap"],
            row["void_patch_count"],
            row["void_largest_patch"],
            row["void_mean_patch"],
            row["roughness_rms"],
            row["roughness_mean_absolute"],
            row["roughness_peak_to_valley"],
            row["primary_name"],
            row["primary_mean"],
            row["secondary_name"],
            row["secondary_mean"],
            row["grid"],
            row["radius_scale"],
            row["hydrogen_included"],
            row["balance_qc"],
            row["union_qc"],
            row["status"],
            row["trajectory"],
            row["summary_path"],
            row["run_directory"],
            row["near_surface"],
            row["anchor_conditioned"],
            row["near_surface_void_largest_patch"],
            row["p_near_surface_void_largest_patch"],
            "GEOMETRIC_MORPHOLOGY_ONLY",
        ]
        for row in results
    ]
    _style_table_sheet(
        results_sheet,
        RESULT_HEADERS,
        result_rows,
        "CoverageResultsTable",
        api,
    )

    for column in (
        "F",
        "G",
        "H",
        "I",
        "J",
        "L",
        "M",
        "N",
        "O",
        "P",
        "R",
        "T",
        "U",
        "V",
        "X",
        "Y",
        "AD",
        "AE",
        "AF",
        "AG",
    ):
        for cell in results_sheet[column][1:]:
            cell.number_format = "0.00"
    for column in ("C", "K"):
        for cell in results_sheet[column][1:]:
            cell.number_format = "0"
    for cell in results_sheet["E"][1:]:
        cell.number_format = "#,##0"
    widths = {
        "A": 42,
        "B": 34,
        "C": 14,
        "D": 16,
        "E": 16,
        "F": 19,
        "G": 15,
        "H": 14,
        "I": 15,
        "J": 21,
        "K": 17,
        "L": 24,
        "M": 24,
        "N": 20,
        "O": 27,
        "P": 28,
        "Q": 24,
        "R": 20,
        "S": 24,
        "T": 22,
        "U": 17,
        "V": 14,
        "W": 18,
        "X": 16,
        "Y": 15,
        "Z": 11,
        "AA": 48,
        "AB": 48,
        "AC": 42,
        "AD": 24,
        "AE": 30,
        "AF": 34,
        "AG": 39,
        "AH": 34,
    }
    for column, width in widths.items():
        results_sheet.column_dimensions[column].width = width

    last_result_row = len(results) + 1
    if results:
        results_sheet.conditional_formatting.add(
            f"F2:F{last_result_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
        results_sheet.conditional_formatting.add(
            f"Z2:Z{last_result_row}",
            FormulaRule(
                formula=["$Z2=\"OK\""],
                fill=PatternFill("solid", fgColor="C6EFCE"),
                font=Font(color="006100"),
            ),
        )
        results_sheet.conditional_formatting.add(
            f"Z2:Z{last_result_row}",
            FormulaRule(
                formula=["$Z2=\"CHECK\""],
                fill=PatternFill("solid", fgColor="FFC7CE"),
                font=Font(color="9C0006"),
            ),
        )

        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Projected coverage by system and temperature"
        chart.x_axis.title = "Total coverage (%)"
        chart.y_axis.title = "System | Temperature"
        chart.height = max(7.0, min(16.0, 2.5 + 0.45 * len(results)))
        chart.width = 17.0
        data = Reference(results_sheet, min_col=6, min_row=1, max_row=last_result_row)
        categories = Reference(results_sheet, min_col=1, min_row=2, max_row=last_result_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        results_sheet.add_chart(chart, "X2")

    component_rows = [
        [
            row["run"],
            row["system"],
            row["temperature"],
            row["name"],
            row["lo"],
            row["hi"],
            row["mean"],
            row["sem"],
            row["std"],
        ]
        for row in components
    ]
    _style_table_sheet(
        component_sheet,
        COMPONENT_HEADERS,
        component_rows,
        "CoverageComponentsTable",
        api,
    )
    for column in ("G", "H", "I"):
        for cell in component_sheet[column][1:]:
            cell.number_format = "0.00"
    for column in ("C", "E", "F"):
        for cell in component_sheet[column][1:]:
            cell.number_format = "0"
    for column, width in {
        "A": 42,
        "B": 34,
        "C": 14,
        "D": 26,
        "E": 19,
        "F": 17,
        "G": 16,
        "H": 15,
        "I": 14,
    }.items():
        component_sheet.column_dimensions[column].width = width

    methods_sheet.sheet_view.showGridLines = False
    methods_sheet.append(["Field", "Value"])
    methods_sheet.append(["Workbook purpose", "Consolidated projected coverage results from completed 300 K and 400 K hold analyses"])
    methods_sheet.append(["Coverage definition", "Total/Uncovered/Component Overlap are a z-UNRESTRICTED canopy metric: periodic union of projected elemental van der Waals disks divided by instantaneous x/y cell area, regardless of how far above the surface an atom sits. A ligand clump that has lifted off and drifted upward with a retracting wall still reads as covered here -- see Near-Surface/Anchor-Conditioned Coverage for a height-gated alternative."])
    methods_sheet.append(["Near-Surface Coverage (%)", "Same projected-union method restricted to ligand atoms within --near-surface-height (default 5 Å) of the local substrate-only height reference; an atom-level gate."])
    methods_sheet.append(["P-Near-Surface-Conditioned Coverage (%)", "Projected union of the FULL footprint of every ligand molecule that has at least one phosphorus atom within the same near-surface height gate. This is a geometric proximity condition, not proof of a Ni-O-P bond."])
    methods_sheet.append(["Scientific Scope", "GEOMETRIC_MORPHOLOGY_ONLY: TECHNICAL Status=OK means arithmetic/file checks passed; it does not validate surface adsorption chemistry or the ligand/NiO cross interaction."])
    methods_sheet.append(["Uncertainty", "Block SEM from the requested trajectory blocks; frame SD is also retained"])
    methods_sheet.append(["Void patches", "Periodic four-connected uncovered regions on the projected-coverage grid; largest and mean areas are percentages of the full x/y cell"])
    methods_sheet.append(["Roughness", "Top-of-film max-z height map over all atoms; RMS, mean-absolute, and peak-to-valley values are experimental morphology proxies"])
    methods_sheet.append(["QC Balance", "Total coverage + uncovered coverage - 100%; expected to be approximately zero"])
    methods_sheet.append(["QC Union", "For one or two components: total - sum(component coverage) + overlap; expected to be approximately zero"])
    methods_sheet.append(["Source root", str(Path(prepared_root).resolve())])
    methods_sheet.append(["Runs included", len(results)])
    for cell in methods_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    methods_sheet.column_dimensions["A"].width = 24
    methods_sheet.column_dimensions["B"].width = 110
    methods_sheet.freeze_panes = "A2"
    for row in methods_sheet.iter_rows():
        row[1].alignment = api["Alignment"](wrap_text=True, vertical="top")

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
