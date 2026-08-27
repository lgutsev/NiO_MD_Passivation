from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path
from typing import Iterable


ATTENTION_STATES = {
    "FINAL_UNMARKED",
    "HASH_MISMATCH",
    "MARKER_WITHOUT_FINAL",
    "MISSING_DIRECTORY",
}

ERROR_DETAIL_DEFAULT = "audit did not run to completion for this campaign"
DEFAULT_AUDIT_TIMEOUT_SECONDS = 300.0
MAX_DETAIL_CHARS = 4000

STAGE_FILES = [
    ("INITIAL_OPT", "xtb_initial_opt.out", "xtb_initial_opt.err", "xtbopt_initial.xyz"),
    ("STEERED_MD", "xtb_steered_md.out", "xtb_steered_md.err", "xtbsteered_last.xyz"),
    ("UNBIASED_MD", "xtb_unbiased_md.out", "xtb_unbiased_md.err", "xtbmd_unbiased_last.xyz"),
    ("FINAL_OPT", "xtb_final_opt.out", "xtb_final_opt.err", "xtbfinal.xyz"),
]

AUDITED_ARTIFACTS = [
    "input.xyz",
    "xtb_input.sha256",
    "xtb_input.complete",
    "xtb_protocol.sha256",
    "xtb_stages.protocol",
    "xtb_protocol.complete",
    "xtbopt_initial.xyz",
    "steering_plan.json",
    "steering_restraints.json",
    "xtb_steered.trj",
    "xtbsteered_last.xyz",
    "xtb_unbiased.trj",
    "xtbmd_unbiased_last.xyz",
    "xtbfinal.xyz",
    "xtbfinal.log",
]

SCHEDULER_FIELDS = [
    "job_id",
    "job_name",
    "partition",
    "state",
    "exit_code",
    "elapsed",
    "time_limit",
    "cpus",
    "node_list",
    "reason",
    "submit_time",
    "start_time",
    "end_time",
]


def _as_int(value: object, default: int = 0) -> int:
    try:
        return int(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return default


def _truncate(text: str, limit: int = MAX_DETAIL_CHARS) -> str:
    text = text.strip()
    if len(text) <= limit:
        return text
    return text[:limit] + f" ... [truncated, {len(text) - limit} more chars]"


def _utc_mtime(path: Path) -> str:
    if not path.exists():
        return ""
    return datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat()


def _sha256(path: Path) -> str:
    if not path.is_file():
        return ""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _same_content(left: Path, right: Path) -> bool | None:
    if not left.is_file() or not right.is_file():
        return None
    return left.read_bytes() == right.read_bytes()


def _tail(path: Path, lines: int) -> str:
    if not path.is_file() or path.stat().st_size == 0:
        return ""
    try:
        content = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError as exc:
        return f"<unreadable: {exc}>"
    return "\n".join(content[-lines:])[-8000:]


def _flatten(prefix: str, value: object) -> Iterable[tuple[str, object]]:
    if isinstance(value, dict):
        for key, child in sorted(value.items()):
            name = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten(name, child)
    elif isinstance(value, list):
        yield prefix, json.dumps(value, sort_keys=True)
    else:
        yield prefix, value


def _package_version() -> str:
    try:
        return version("nio-md-prep")
    except PackageNotFoundError:
        return "source-tree"


def _likely_issue(text: str, scheduler_state: str = "") -> str:
    combined = f"{scheduler_state}\n{text}".lower()
    signatures = [
        (("out_of_memory", "oom", "cannot allocate memory"), "OUT_OF_MEMORY"),
        (("time limit", "timeout"), "WALLTIME_EXCEEDED"),
        (("cancelled", "canceled"), "CANCELLED"),
        (("segmentation fault", "segfault"), "SEGMENTATION_FAULT"),
        (("killed", "terminated by signal"), "PROCESS_KILLED"),
        (("scf not converged", "failed to converge"), "SCF_NOT_CONVERGED"),
        (("command not found", "xtb not found"), "EXECUTABLE_NOT_FOUND"),
        (("no space left",), "FILESYSTEM_FULL"),
        (("permission denied",), "PERMISSION_ERROR"),
        (("failed", "error"), "APPLICATION_ERROR"),
    ]
    for needles, diagnosis in signatures:
        if any(needle in combined for needle in needles):
            return diagnosis
    return "NO_EXPLICIT_ERROR_SIGNATURE"


def _next_action(status: str) -> str:
    return {
        "COMPLETE": "ready for VASP regeneration",
        "PENDING": "run initial optimization",
        "INITIAL_OPT_STARTED": "resume/check initial optimization",
        "INITIAL_OPT_COMPLETE": "run steered or unbiased MD",
        "STEERING_COMPLETE": "run unbiased MD, then final optimization",
        "UNBIASED_MD_COMPLETE": "run final optimization",
        "FINAL_UNMARKED": "validate final geometry and create completion marker",
        "HASH_MISMATCH": "validate/migrate legacy result with --regenerate-vasp",
        "MARKER_WITHOUT_FINAL": "rerun final optimization",
        "MISSING_DIRECTORY": "restore or regenerate case directory",
    }.get(status, "inspect case logs and resume")


def _log_attempts(campaign: Path) -> tuple[list[dict[str, object]], dict[int, dict]]:
    rows: list[dict[str, object]] = []
    attempts: dict[tuple[str, int], dict[str, object]] = {}
    array_pattern = re.compile(r"^xtb\.(?P<job>\d+)_(?P<index>\d+)\.(?P<stream>out|err)$")
    pool_pattern = re.compile(r"^xtb-pool\.(?P<job>\d+)\.(?P<stream>out|err)$")
    start_pattern = re.compile(r">>> Starting xTB case index (\d+) at (.*)")
    finish_pattern = re.compile(
        r"<<< Finished xTB case index (\d+) (OK|with ERROR (\d+)) at (.*)"
    )
    for path in sorted(campaign.glob("xtb*.out")) + sorted(campaign.glob("xtb*.err")):
        match = array_pattern.match(path.name)
        if match:
            job_id, index = match.group("job"), int(match.group("index"))
            attempt = attempts.setdefault((job_id, index), {})
            attempt.update(job_id=job_id, array_index=index)
            attempt[f"{match.group('stream')}_file"] = path.name
            attempt[f"{match.group('stream')}_tail"] = _tail(path, 30)
            attempt["last_modified_utc"] = max(
                str(attempt.get("last_modified_utc", "")), _utc_mtime(path)
            )
            continue
        match = pool_pattern.match(path.name)
        if not match:
            continue
        job_id = match.group("job")
        text = path.read_text(encoding="utf-8", errors="replace")
        for line in text.splitlines():
            start = start_pattern.search(line)
            if start:
                index = int(start.group(1))
                attempt = attempts.setdefault((job_id, index), {})
                attempt.update(job_id=job_id, array_index=index, started_at=start.group(2))
            finish = finish_pattern.search(line)
            if finish:
                index = int(finish.group(1))
                attempt = attempts.setdefault((job_id, index), {})
                attempt.update(
                    job_id=job_id,
                    array_index=index,
                    worker_state="COMPLETED" if finish.group(2) == "OK" else "FAILED",
                    worker_exit_code="0" if finish.group(2) == "OK" else finish.group(3),
                    finished_at=finish.group(4),
                )
        for (attempt_job, _), attempt in attempts.items():
            if attempt_job == job_id:
                attempt[f"pool_{match.group('stream')}_file"] = path.name
                attempt[f"pool_{match.group('stream')}_tail"] = _tail(path, 30)
                attempt["last_modified_utc"] = max(
                    str(attempt.get("last_modified_utc", "")), _utc_mtime(path)
                )
    by_index: dict[int, dict] = {}
    for (_, index), attempt in attempts.items():
        rows.append(dict(attempt))
        previous = by_index.get(index)
        if previous is None or str(attempt.get("last_modified_utc", "")) >= str(
            previous.get("last_modified_utc", "")
        ):
            by_index[index] = attempt
    return rows, by_index


def _query_sacct(job_ids: Iterable[str]) -> list[dict[str, object]]:
    unique = sorted({job_id for job_id in job_ids if job_id})
    if not unique or shutil.which("sacct") is None:
        return []
    command = [
        "sacct", "-n", "-P", "-j", ",".join(unique),
        "--format=JobIDRaw,JobName,Partition,State,ExitCode,Elapsed,Timelimit,NCPUS,NodeList,Reason,Submit,Start,End",
    ]
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=20, check=False)
    except (OSError, subprocess.TimeoutExpired):
        return []
    if result.returncode:
        return []
    records = []
    for line in result.stdout.splitlines():
        fields = line.split("|")
        if len(fields) < len(SCHEDULER_FIELDS):
            continue
        records.append(dict(zip(SCHEDULER_FIELDS, fields[: len(SCHEDULER_FIELDS)])))
    return records


def _campaign_manifests(root: Path) -> tuple[list[tuple[Path, dict]], list[str]]:
    candidates: set[Path] = set()
    warnings: list[str] = []

    def on_error(exc: OSError) -> None:
        warnings.append(f"could not scan {exc.filename}: {exc.strerror or exc}")

    for dirpath, _dirnames, filenames in os.walk(root, onerror=on_error):
        current = Path(dirpath)
        if "agglomeration_manifest.json" in filenames or current.name.endswith("_agglo"):
            candidates.add(current)
    if root.name.endswith("_agglo"):
        candidates.add(root)
    campaigns: list[tuple[Path, dict]] = []
    for campaign in sorted(candidates):
        manifest_path = campaign / "agglomeration_manifest.json"
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except OSError:
            campaigns.append((campaign, {"_manifest_error": "manifest is missing"}))
            continue
        except UnicodeDecodeError as exc:
            campaigns.append(
                (campaign, {"_manifest_error": f"manifest is not valid UTF-8: {exc}"})
            )
            continue
        except json.JSONDecodeError as exc:
            campaigns.append(
                (campaign, {"_manifest_error": f"manifest is invalid JSON: {exc}"})
            )
            continue
        if isinstance(manifest, dict) and isinstance(manifest.get("replicas"), list):
            campaigns.append((campaign, manifest))
        elif campaign.name.endswith("_agglo"):
            campaigns.append(
                (campaign, {"_manifest_error": "manifest lacks campaign replicas"})
            )
    return campaigns, warnings


def _cross_check_progress(progress: dict) -> str | None:
    cases = progress.get("cases")
    if not isinstance(cases, list) or not cases:
        return None
    actual = Counter(
        str(case.get("status", "UNKNOWN"))
        for case in cases
        if isinstance(case, dict)
    )
    reported = progress.get("counts_by_status", {})
    if not isinstance(reported, dict):
        reported = {}
    normalized = {str(key): _as_int(value) for key, value in reported.items()}
    if dict(actual) != normalized:
        return (
            "counts_by_status does not match per-case rows "
            f"(reported {normalized}, computed {dict(actual)})"
        )
    if progress.get("total") is not None and _as_int(progress["total"], -1) != len(cases):
        return f"total ({progress['total']}) does not match case rows ({len(cases)})"
    return None


def _campaign_label(root: Path, campaign: Path) -> str:
    try:
        relative = campaign.relative_to(root)
    except ValueError:
        return str(campaign)
    return "." if str(relative) == "." else str(relative)


def _case_diagnostics(
    campaign_label: str,
    campaign: Path,
    audited_cases: list[dict],
    latest_attempts: dict[int, dict],
    scheduler_by_job: dict[str, dict],
    log_tail_lines: int,
) -> tuple[list[dict], list[dict], list[dict]]:
    cases: list[dict] = []
    artifacts: list[dict] = []
    logs: list[dict] = []
    for source in audited_cases:
        array_index = int(source.get("array_index", -1))
        relative = str(source.get("case", ""))
        work = campaign / relative
        attempt = latest_attempts.get(array_index, {})
        job_id = str(attempt.get("job_id", ""))
        scheduler = scheduler_by_job.get(job_id, {})
        stage_values: dict[str, bool] = {}
        existing_logs: list[Path] = []
        for stage, out_name, err_name, checkpoint_name in STAGE_FILES:
            stage_values[f"{stage.lower()}_complete"] = (work / checkpoint_name).is_file()
            for stream, name in (("stdout", out_name), ("stderr", err_name)):
                path = work / name
                if not path.is_file():
                    continue
                existing_logs.append(path)
                tail = _tail(path, log_tail_lines)
                logs.append(
                    {
                        "campaign": campaign_label,
                        "array_index": array_index,
                        "case": relative,
                        "stage": stage,
                        "stream": stream,
                        "file": name,
                        "bytes": path.stat().st_size,
                        "modified_utc": _utc_mtime(path),
                        "sha256": _sha256(path),
                        "tail": tail,
                        "error_signature": _likely_issue(tail),
                    }
                )
        for name in AUDITED_ARTIFACTS:
            path = work / name
            artifacts.append(
                {
                    "campaign": campaign_label,
                    "array_index": array_index,
                    "case": relative,
                    "artifact": name,
                    "present": path.is_file(),
                    "bytes": path.stat().st_size if path.is_file() else 0,
                    "modified_utc": _utc_mtime(path),
                    "sha256": _sha256(path),
                }
            )
        nonempty_logs = [path for path in existing_logs if path.stat().st_size > 0]
        latest_log = max(
            nonempty_logs or existing_logs,
            key=lambda item: item.stat().st_mtime,
            default=None,
        )
        latest_tail = _tail(latest_log, log_tail_lines) if latest_log else ""
        pool_tail = "\n".join(
            str(attempt.get(key, ""))
            for key in ("pool_err_tail", "pool_out_tail", "err_tail", "out_tail")
            if attempt.get(key)
        )
        scheduler_state = str(scheduler.get("state", attempt.get("worker_state", "")))
        diagnosis_text = "\n".join(part for part in (latest_tail, pool_tail) if part)
        status = str(source.get("status", "UNKNOWN"))
        case_row = {
            "campaign": campaign_label,
            "array_index": array_index,
            "case": relative,
            "work_directory": str(work.resolve()),
            "status": status,
            "safe_complete": status == "COMPLETE",
            "charge": source.get("charge", ""),
            "uhf": source.get("uhf", ""),
            "detail": source.get("detail", ""),
            "next_action": _next_action(status),
            "likely_issue": (
                "NONE" if status == "COMPLETE" else _likely_issue(diagnosis_text, scheduler_state)
            ),
            "latest_log": latest_log.name if latest_log else "",
            "latest_log_modified_utc": _utc_mtime(latest_log) if latest_log else "",
            "latest_log_tail": latest_tail,
            "input_hash_match": _same_content(
                work / "xtb_input.sha256", work / "xtb_input.complete"
            ),
            "protocol_stage_hash_match": _same_content(
                work / "xtb_protocol.sha256", work / "xtb_stages.protocol"
            ),
            "protocol_completion_hash_match": _same_content(
                work / "xtb_protocol.sha256", work / "xtb_protocol.complete"
            ),
            "latest_job_id": job_id,
            "worker_state": attempt.get("worker_state", ""),
            "worker_exit_code": attempt.get("worker_exit_code", ""),
            "slurm_state": scheduler_state,
            "slurm_exit_code": scheduler.get("exit_code", ""),
            "slurm_reason": scheduler.get("reason", ""),
            "slurm_elapsed": scheduler.get("elapsed", ""),
            "slurm_time_limit": scheduler.get("time_limit", ""),
            "attempt_started_at": attempt.get("started_at", scheduler.get("start_time", "")),
            "attempt_finished_at": attempt.get("finished_at", scheduler.get("end_time", "")),
            **stage_values,
        }
        cases.append(case_row)
    return cases, artifacts, logs


def _campaign_protocol_rows(campaign_label: str, manifest: dict) -> list[dict]:
    rows = []
    for section in ("xtb", "vasp_md", "agglomeration"):
        for key, value in _flatten("", manifest.get(section, {})):
            rows.append(
                {
                    "campaign": campaign_label,
                    "section": section,
                    "key": key,
                    "value": value,
                    "value_type": type(value).__name__,
                }
            )
    return rows


def _campaign_resource_row(campaign_label: str, campaign: Path, manifest: dict) -> dict:
    settings = manifest.get("xtb", {}) if isinstance(manifest.get("xtb"), dict) else {}
    pool = campaign / "run_xtb_pool.sbatch"
    array = campaign / "run_xtb_array.sbatch"
    manifest_path = campaign / "agglomeration_manifest.json"
    case_list = campaign / "xtb_cases.tsv"
    audit = campaign / "audit_xtb_runs.py"
    launcher = campaign / "launch_xtb.sh"
    return {
        "campaign": campaign_label,
        "partition": settings.get("partition", ""),
        "account": settings.get("account", ""),
        "cpus_per_case": settings.get("cpus", ""),
        "walltime": settings.get("walltime", ""),
        "array_concurrency": settings.get("array_concurrency", ""),
        "node_pool_enabled": settings.get("node_pool_enabled", ""),
        "node_cpus": settings.get("node_cpus", ""),
        "pool_workers": settings.get("pool_workers", ""),
        "omp_stacksize": settings.get("omp_stacksize", ""),
        "pool_script_present": pool.is_file(),
        "pool_script_sha256": _sha256(pool),
        "array_script_present": array.is_file(),
        "array_script_sha256": _sha256(array),
        "manifest_sha256": _sha256(manifest_path),
        "case_list_sha256": _sha256(case_list),
        "audit_script_sha256": _sha256(audit),
        "launcher_sha256": _sha256(launcher),
    }


def _audit_campaign(
    root: Path,
    campaign: Path,
    manifest: dict,
    *,
    include_slurm: bool,
    log_tail_lines: int,
    timeout: float,
) -> dict[str, object]:
    row: dict[str, object] = {
        "campaign": _campaign_label(root, campaign),
        "path": str(campaign.resolve()),
        "campaign_status": "ERROR",
        "manifest_status": str(manifest.get("status", "UNKNOWN")),
        "total": 0,
        "completed": 0,
        "pending": 0,
        "attention": 0,
        "percent_complete": 0.0,
        "counts_by_status": {},
        "data_consistent": True,
        "audit_seconds": None,
        "detail": ERROR_DETAIL_DEFAULT,
    }
    if manifest.get("_manifest_error"):
        row["detail"] = str(manifest["_manifest_error"])
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": [], "resources": [], "slurm_jobs": []}
    case_list = campaign / "xtb_cases.tsv"
    audit = campaign / "audit_xtb_runs.py"
    if not case_list.is_file():
        if manifest.get("status") == "PACKMOL_REQUIRED":
            row["campaign_status"] = "PACKMOL_REQUIRED"
            row["detail"] = "Packmol structures must be generated before xTB"
        elif not manifest.get("xtb", {}).get("enabled", False):
            row["campaign_status"] = "XTB_DISABLED"
            row["detail"] = "xTB is disabled for this campaign"
        else:
            row["detail"] = "xtb_cases.tsv is missing"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    if not audit.is_file():
        row["detail"] = (
            "audit_xtb_runs.py is missing; regenerate this campaign's xTB launcher"
        )
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}

    started = time.monotonic()
    try:
        result = subprocess.run(
            [sys.executable, str(audit), "--summary-only"],
            cwd=campaign,
            capture_output=True,
            text=True,
            check=False,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        row["audit_seconds"] = round(time.monotonic() - started, 2)
        row["detail"] = f"audit_xtb_runs.py did not finish within {timeout:.0f}s"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    except OSError as exc:
        row["detail"] = f"could not launch audit_xtb_runs.py: {exc}"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    row["audit_seconds"] = round(time.monotonic() - started, 2)
    if result.returncode:
        row["detail"] = _truncate(
            f"audit failed with exit code {result.returncode}: "
            f"{result.stderr or result.stdout}"
        )
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    progress_path = campaign / "xtb_progress.json"
    try:
        progress_text = progress_path.read_text(encoding="utf-8")
    except OSError as exc:
        row["detail"] = f"audit did not produce readable xtb_progress.json: {exc}"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    try:
        progress = json.loads(progress_text)
    except json.JSONDecodeError as exc:
        row["detail"] = f"xtb_progress.json is not valid JSON: {exc}"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}
    if not isinstance(progress, dict):
        row["detail"] = "xtb_progress.json did not contain a JSON object"
        return {"summary": row, "cases": [], "artifacts": [], "logs": [], "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest), "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)], "slurm_jobs": []}

    inconsistency = _cross_check_progress(progress)
    counts = progress.get("counts_by_status", {})
    if not isinstance(counts, dict):
        counts = {}
    total = _as_int(progress.get("total"))
    completed = _as_int(progress.get("completed"))
    pending = _as_int(progress.get("pending"), max(0, total - completed))
    attention = sum(_as_int(counts.get(status)) for status in ATTENTION_STATES)
    row.update(
        total=total,
        completed=completed,
        pending=pending,
        attention=attention,
        percent_complete=round(100.0 if total == 0 else 100.0 * completed / total, 1),
        counts_by_status=dict(sorted(counts.items())),
    )
    if inconsistency:
        row["campaign_status"] = "ATTENTION"
        row["data_consistent"] = False
        row["detail"] = f"inconsistent audit output: {inconsistency}"
    elif attention:
        row["campaign_status"] = "ATTENTION"
        row["detail"] = "one or more cases have inconsistent/missing completion data"
    elif pending:
        row["campaign_status"] = "IN_PROGRESS"
        row["detail"] = "xTB work remains"
    else:
        row["campaign_status"] = "COMPLETE"
        row["detail"] = "all xTB cases are safely complete"
    attempt_rows, latest_attempts = _log_attempts(campaign)
    slurm_jobs = _query_sacct(str(item.get("job_id", "")) for item in attempt_rows) if include_slurm else []
    scheduler_by_job = {
        str(job.get("job_id", "")).split(".", 1)[0]: job
        for job in slurm_jobs
        if "." not in str(job.get("job_id", ""))
    }
    case_rows, artifact_rows, log_rows = _case_diagnostics(
        str(row["campaign"]),
        campaign,
        list(progress.get("cases", [])),
        latest_attempts,
        scheduler_by_job,
        log_tail_lines,
    )
    for attempt in attempt_rows:
        attempt["campaign"] = row["campaign"]
        attempt["record_type"] = "LOG_ATTEMPT"
    for job in slurm_jobs:
        job["campaign"] = row["campaign"]
        job["record_type"] = "SACCT"
    return {
        "summary": row,
        "cases": case_rows,
        "artifacts": artifact_rows,
        "logs": log_rows,
        "protocols": _campaign_protocol_rows(str(row["campaign"]), manifest),
        "resources": [_campaign_resource_row(str(row["campaign"]), campaign, manifest)],
        "slurm_jobs": attempt_rows + slurm_jobs,
    }


def _write_csv(path: Path, rows: list[dict], preferred_fields: list[str] | None = None) -> None:
    fields = list(preferred_fields or [])
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        if not fields:
            handle.write("")
            return
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            serialized = {
                key: json.dumps(value, sort_keys=True)
                if isinstance(value, (dict, list))
                else value
                for key, value in row.items()
            }
            writer.writerow(serialized)


def _excel_value(value: object) -> object:
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True)
    if value is None:
        return ""
    text = str(value)
    return text[:32767] if len(text) > 32767 else value


def _write_workbook(path: Path, report: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    workbook.remove(workbook.active)
    dark = "17324D"
    blue = "2F75B5"
    green = "C6EFCE"
    yellow = "FFF2CC"
    red = "F4CCCC"
    gray = "E7E6E6"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E2F3")

    def add_table_sheet(name: str, rows: list[dict], fields: list[str]) -> object:
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.append(fields)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            sheet.append([_excel_value(row.get(field, "")) for field in fields])
        if rows:
            table_name = re.sub(r"[^A-Za-z0-9_]", "", name) + "Table"
            table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(fields))}{len(rows)+1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_index, field in enumerate(fields, 1):
            sample = [str(row.get(field, "")) for row in rows[:200]]
            width = min(60, max(10, len(field) + 2, *(min(60, len(value) + 2) for value in sample)))
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.auto_filter.ref = sheet.dimensions
        return sheet

    overview = workbook.create_sheet("Overview")
    overview.sheet_view.showGridLines = False
    overview["A1"] = "Agglomeration xTB Audit"
    overview["A1"].font = Font(size=18, bold=True, color=white)
    overview["A1"].fill = PatternFill("solid", fgColor=dark)
    overview.merge_cells("A1:F1")
    overview["A2"] = "Audited at (UTC)"
    overview["B2"] = report["audited_at_utc"]
    overview["A3"] = "Root"
    overview["B3"] = report["root"]
    overview.merge_cells("B3:F3")
    kpis = [
        ("Campaigns", report["campaigns"]),
        ("Cases", report["cases"]),
        ("Safely complete", report["completed"]),
        ("Pending/attention", report["pending"]),
        ("Attention cases", report["attention"]),
    ]
    for column, (label, value) in enumerate(kpis, 1):
        cell = overview.cell(5, column, label)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        value_cell = overview.cell(6, column, value)
        value_cell.font = Font(size=16, bold=True)
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = Border(bottom=thin)
    overview["A8"] = "Campaign"
    overview["B8"] = "Complete"
    overview["C8"] = "Pending"
    for cell in overview[8][:3]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color=white, bold=True)
    for offset, row in enumerate(report["campaign_results"], 9):
        overview.cell(offset, 1, row["campaign"])
        overview.cell(offset, 2, row["completed"])
        overview.cell(offset, 3, row["pending"])
    if report["campaign_results"]:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "xTB Case Progress by Campaign"
        chart.y_axis.title = "Campaign"
        chart.x_axis.title = "Cases"
        chart.add_data(
            Reference(overview, min_col=2, max_col=3, min_row=8, max_row=8 + len(report["campaign_results"])),
            titles_from_data=True,
        )
        chart.series[0].tx = SeriesLabel(v="Complete")
        chart.series[1].tx = SeriesLabel(v="Pending")
        chart.set_categories(
            Reference(overview, min_col=1, min_row=9, max_row=8 + len(report["campaign_results"]))
        )
        chart.height = 8
        chart.width = 16
        chart.x_axis.majorUnit = 1
        chart.x_axis.numFmt = "0"
        overview.add_chart(chart, "H2")
    for column in "ABCDEF":
        overview.column_dimensions[column].width = 22

    campaign_fields = [
        "campaign", "path", "campaign_status", "manifest_status", "total",
        "completed", "pending", "attention", "percent_complete",
        "counts_by_status", "data_consistent", "audit_seconds", "detail",
    ]
    incomplete_case_fields = [
        "campaign", "array_index", "case", "status", "detail", "likely_issue",
        "next_action", "latest_log", "latest_log_modified_utc", "latest_job_id",
        "slurm_state", "slurm_exit_code", "slurm_reason", "slurm_elapsed",
        "slurm_time_limit",
    ]
    case_fields = [
        "campaign", "array_index", "case", "work_directory", "status",
        "safe_complete", "charge", "uhf", "detail", "next_action",
        "likely_issue", "latest_log", "latest_log_modified_utc", "latest_log_tail",
        "input_hash_match", "protocol_stage_hash_match",
        "protocol_completion_hash_match", "latest_job_id", "worker_state",
        "worker_exit_code", "slurm_state", "slurm_exit_code", "slurm_reason",
        "slurm_elapsed", "slurm_time_limit", "attempt_started_at",
        "attempt_finished_at", "initial_opt_complete", "steered_md_complete",
        "unbiased_md_complete", "final_opt_complete",
    ]
    artifact_fields = [
        "campaign", "array_index", "case", "artifact", "present", "bytes",
        "modified_utc", "sha256",
    ]
    log_fields = [
        "campaign", "array_index", "case", "stage", "stream", "file", "bytes",
        "modified_utc", "sha256", "error_signature", "tail",
    ]
    protocol_fields = ["campaign", "section", "key", "value", "value_type"]
    resource_fields = [
        "campaign", "partition", "account", "cpus_per_case", "walltime",
        "array_concurrency", "node_pool_enabled", "node_cpus", "pool_workers",
        "omp_stacksize", "pool_script_present", "pool_script_sha256",
        "array_script_present", "array_script_sha256",
        "manifest_sha256", "case_list_sha256", "audit_script_sha256",
        "launcher_sha256",
    ]
    slurm_fields = [
        "campaign", "record_type", "job_id", "array_index", "worker_state",
        "worker_exit_code", "started_at", "finished_at", "last_modified_utc",
        *[field for field in SCHEDULER_FIELDS if field != "job_id"],
        "out_file", "err_file", "pool_out_file", "pool_err_file",
        "out_tail", "err_tail", "pool_out_tail", "pool_err_tail",
    ]
    campaign_sheet = add_table_sheet("Campaigns", report["campaign_results"], campaign_fields)
    incomplete_sheet = add_table_sheet(
        "Incomplete Cases", report["incomplete_cases"], incomplete_case_fields
    )
    all_cases_sheet = add_table_sheet("All Cases", report["case_results"], case_fields)
    integrity_fields = [
        "campaign", "array_index", "case", "work_directory", "status",
        "initial_opt_complete", "steered_md_complete", "unbiased_md_complete",
        "final_opt_complete", "input_hash_match", "protocol_stage_hash_match",
        "protocol_completion_hash_match",
    ]
    add_table_sheet("Case Integrity", report["case_results"], integrity_fields)
    add_table_sheet("Artifacts", report["artifact_results"], artifact_fields)
    add_table_sheet("Logs", report["log_results"], log_fields)
    add_table_sheet("Protocols", report["protocol_results"], protocol_fields)
    add_table_sheet("Resources", report["resource_results"], resource_fields)
    add_table_sheet("Slurm Jobs", report["slurm_jobs"], slurm_fields)
    metadata = [
        {"key": "schema_version", "value": report["schema_version"], "meaning": "Audit report schema"},
        {"key": "generated_by", "value": report["generated_by"], "meaning": "Installed package version"},
        {"key": "python_version", "value": report["python_version"], "meaning": "Python interpreter"},
        {"key": "hostname", "value": report["hostname"], "meaning": "Host that performed audit"},
        {"key": "audited_at_utc", "value": report["audited_at_utc"], "meaning": "Audit timestamp"},
        {"key": "root", "value": report["root"], "meaning": "Audited campaign root"},
        {"key": "slurm_correlation_requested", "value": report["slurm_correlation_requested"], "meaning": "Whether sacct correlation was enabled"},
        {"key": "sacct_available", "value": report["sacct_available"], "meaning": "Whether sacct was found on PATH"},
        {"key": "log_tail_lines", "value": report["log_tail_lines"], "meaning": "Retained lines per stage log"},
        {"key": "audit_timeout_seconds", "value": report["audit_timeout_seconds"], "meaning": "Maximum runtime allowed per campaign audit"},
        {"key": "scan_warning_count", "value": len(report["scan_warnings"]), "meaning": "Directories that could not be scanned"},
        {"key": "scan_warnings", "value": report["scan_warnings"], "meaning": "Filesystem traversal warnings"},
    ]
    add_table_sheet("Audit Metadata", metadata, ["key", "value", "meaning"])
    definitions = [
        {"status": "COMPLETE", "meaning": "Final geometry and matching completion hash are present", "recommended_action": "Regenerate VASP inputs"},
        {"status": "INITIAL_OPT_COMPLETE", "meaning": "Initial optimization checkpoint exists", "recommended_action": "Resume MD stages"},
        {"status": "INITIAL_OPT_STARTED", "meaning": "Initial optimization output exists without its checkpoint", "recommended_action": "Inspect log and resume initial optimization"},
        {"status": "PENDING", "meaning": "No completed stage checkpoint exists", "recommended_action": "Run initial optimization"},
        {"status": "STEERING_COMPLETE", "meaning": "Steered MD checkpoint exists", "recommended_action": "Resume unbiased MD and final optimization"},
        {"status": "UNBIASED_MD_COMPLETE", "meaning": "Unbiased MD checkpoint exists", "recommended_action": "Resume final optimization"},
        {"status": "FINAL_UNMARKED", "meaning": "Final XYZ exists without completion marker", "recommended_action": "Validate before marking complete"},
        {"status": "HASH_MISMATCH", "meaning": "Final XYZ completion marker is for another protocol", "recommended_action": "Validate/migrate legacy result"},
        {"status": "MARKER_WITHOUT_FINAL", "meaning": "Completion marker exists but final XYZ is absent", "recommended_action": "Rerun final optimization"},
        {"status": "MISSING_DIRECTORY", "meaning": "TSV case directory is absent", "recommended_action": "Restore or regenerate directory"},
        {"status": "NO_EXPLICIT_ERROR_SIGNATURE", "meaning": "No known fatal text signature was found", "recommended_action": "Inspect latest log and Slurm record"},
    ]
    add_table_sheet(
        "Status Definitions", definitions, ["status", "meaning", "recommended_action"]
    )

    for sheet in (campaign_sheet, incomplete_sheet, all_cases_sheet):
        status_column = 3 if sheet.title == "Campaigns" else 5
        status_letter = get_column_letter(status_column)
        last_row = max(2, sheet.max_row)
        sheet.conditional_formatting.add(
            f"{status_letter}2:{status_letter}{last_row}",
            FormulaRule(formula=[f'{status_letter}2="COMPLETE"'], fill=PatternFill("solid", fgColor=green)),
        )
        sheet.conditional_formatting.add(
            f"{status_letter}2:{status_letter}{last_row}",
            FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("PENDING",{status_letter}2)),ISNUMBER(SEARCH("PROGRESS",{status_letter}2)))'], fill=PatternFill("solid", fgColor=yellow)),
        )
        sheet.conditional_formatting.add(
            f"{status_letter}2:{status_letter}{last_row}",
            FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("ATTENTION",{status_letter}2)),ISNUMBER(SEARCH("MISMATCH",{status_letter}2)),ISNUMBER(SEARCH("MISSING",{status_letter}2)))'], fill=PatternFill("solid", fgColor=red)),
        )
    overview.freeze_panes = "A5"
    overview.sheet_properties.pageSetUpPr.fitToPage = True
    workbook.save(path)


def audit_agglomerations(
    root: Path,
    output_prefix: Path | None = None,
    *,
    include_slurm: bool = True,
    log_tail_lines: int = 30,
    timeout: float = DEFAULT_AUDIT_TIMEOUT_SECONDS,
) -> dict:
    root = root.resolve()
    if not root.is_dir():
        raise ValueError(f"agglomeration audit root is not a directory: {root}")
    campaigns, scan_warnings = _campaign_manifests(root)
    if not campaigns:
        raise ValueError(f"no agglomeration campaigns found under {root}")

    if log_tail_lines < 1:
        raise ValueError("log_tail_lines must be at least 1")
    if timeout <= 0:
        raise ValueError("timeout must be greater than zero")
    campaign_audits = [
        _audit_campaign(
            root,
            campaign,
            manifest,
            include_slurm=include_slurm,
            log_tail_lines=log_tail_lines,
            timeout=timeout,
        )
        for campaign, manifest in campaigns
    ]
    rows = [audit["summary"] for audit in campaign_audits]
    case_rows = [row for audit in campaign_audits for row in audit["cases"]]
    artifact_rows = [row for audit in campaign_audits for row in audit["artifacts"]]
    log_rows = [row for audit in campaign_audits for row in audit["logs"]]
    protocol_rows = [row for audit in campaign_audits for row in audit["protocols"]]
    resource_rows = [row for audit in campaign_audits for row in audit["resources"]]
    slurm_rows = [row for audit in campaign_audits for row in audit["slurm_jobs"]]
    incomplete_rows = [row for row in case_rows if not row["safe_complete"]]
    attention_rows = [
        row for row in case_rows if str(row.get("status")) in ATTENTION_STATES
    ]
    pending_rows = [
        row for row in case_rows if not row["safe_complete"] and row not in attention_rows
    ]
    status_counts = Counter(str(row["campaign_status"]) for row in rows)
    totals = {
        "campaigns": len(rows),
        "cases": sum(int(row["total"]) for row in rows),
        "completed": sum(int(row["completed"]) for row in rows),
        "pending": sum(int(row["pending"]) for row in rows),
        "attention": sum(int(row["attention"]) for row in rows),
    }
    report = {
        "schema_version": 2,
        "generated_by": f"nio-md-prep {_package_version()}",
        "python_version": sys.version.split()[0],
        "hostname": socket.gethostname(),
        "audited_at_utc": datetime.now(timezone.utc).isoformat(),
        "root": str(root),
        "slurm_correlation_requested": include_slurm,
        "sacct_available": shutil.which("sacct") is not None,
        "log_tail_lines": log_tail_lines,
        "audit_timeout_seconds": timeout,
        "scan_warnings": scan_warnings,
        **totals,
        "campaign_counts_by_status": dict(sorted(status_counts.items())),
        "campaign_results": rows,
        "campaign_issues": [row for row in rows if row["campaign_status"] in {"ERROR", "ATTENTION"}],
        "case_results": case_rows,
        "case_rows": case_rows,
        "incomplete_cases": incomplete_rows,
        "pending_cases": pending_rows,
        "attention_cases": attention_rows,
        "artifact_results": artifact_rows,
        "log_results": log_rows,
        "protocol_results": protocol_rows,
        "resource_results": resource_rows,
        "slurm_jobs": slurm_rows,
    }

    prefix = (output_prefix or (root / "agglomeration_xtb_audit")).resolve()
    prefix.parent.mkdir(parents=True, exist_ok=True)
    csv_path = prefix.with_suffix(".csv")
    json_path = prefix.with_suffix(".json")
    case_csv_path = prefix.with_name(prefix.name + "_cases").with_suffix(".csv")
    incomplete_csv_path = prefix.with_name(prefix.name + "_incomplete").with_suffix(".csv")
    attention_csv_path = prefix.with_name(prefix.name + "_attention").with_suffix(".csv")
    artifact_csv_path = prefix.with_name(prefix.name + "_artifacts").with_suffix(".csv")
    log_csv_path = prefix.with_name(prefix.name + "_logs").with_suffix(".csv")
    protocol_csv_path = prefix.with_name(prefix.name + "_protocols").with_suffix(".csv")
    resource_csv_path = prefix.with_name(prefix.name + "_resources").with_suffix(".csv")
    slurm_csv_path = prefix.with_name(prefix.name + "_slurm").with_suffix(".csv")
    workbook_path = prefix.with_suffix(".xlsx")
    _write_csv(csv_path, rows)
    _write_csv(case_csv_path, case_rows)
    _write_csv(incomplete_csv_path, incomplete_rows)
    _write_csv(attention_csv_path, attention_rows)
    _write_csv(artifact_csv_path, artifact_rows)
    _write_csv(log_csv_path, log_rows)
    _write_csv(protocol_csv_path, protocol_rows)
    _write_csv(resource_csv_path, resource_rows)
    _write_csv(slurm_csv_path, slurm_rows)
    json_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    _write_workbook(workbook_path, report)

    print("Campaign                         Status           Complete   Pending  Attention")
    print("-------------------------------- ---------------- ---------- --------- ---------")
    for row in rows:
        print(
            f"{str(row['campaign'])[:32]:32} "
            f"{str(row['campaign_status']):16} "
            f"{int(row['completed']):4}/{int(row['total']):<5} "
            f"{int(row['pending']):9} {int(row['attention']):9}"
        )
    for warning in scan_warnings:
        print(f"SCAN WARNING: {warning}")
    for row in rows:
        if row["campaign_status"] in {"ATTENTION", "ERROR"}:
            print(f"{row['campaign_status']} {row['campaign']}: {row['detail']}")
    if incomplete_rows:
        print("\nExact incomplete xTB cases:")
        print("Campaign                         Index  Status                    Case")
        print("-------------------------------- ------ ------------------------- ----------------------------------------")
        for case in incomplete_rows:
            print(
                f"{str(case['campaign'])[:32]:32} "
                f"{int(case['array_index']):6} "
                f"{str(case['status'])[:25]:25} "
                f"{case['case']}"
            )
            print(
                f"  diagnosis={case['likely_issue']}; latest_log={case['latest_log'] or 'none'}; "
                f"job={case['latest_job_id'] or 'unknown'}; slurm={case['slurm_state'] or 'unknown'}; "
                f"next={case['next_action']}"
            )
    print(
        f"Overall xTB progress: {totals['completed']}/{totals['cases']} safely complete; "
        f"{totals['pending']} pending or requiring attention across "
        f"{totals['campaigns']} campaign(s)."
    )
    print(
        "Reports written: "
        f"{csv_path}, {case_csv_path}, {incomplete_csv_path}, {attention_csv_path}, "
        f"{artifact_csv_path}, {log_csv_path}, {protocol_csv_path}, "
        f"{resource_csv_path}, {slurm_csv_path}, {json_path}, and {workbook_path}"
    )
    return report
